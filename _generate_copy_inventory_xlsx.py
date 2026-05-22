# -*- coding: utf-8 -*-
"""One-off generator: 游戏文案与操作走查.xlsx"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ASCII filename avoids mojibake on some Windows / toolchains; content is still Chinese.
OUT = Path(__file__).resolve().parent / "copy_inventory_walkthrough.xlsx"


def autosize_columns(ws, max_width=72):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max_width, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = width


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize_columns(ws)
    return ws


def main():
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "说明"
    ws0.append(["字段", "说明"])
    ws0.append(["生成时间", "用于整体走查：页面流程、谜题/密码、OA 检索、分支文案要点、待统一项。"])
    ws0.append(["文件名", OUT.name])
    ws0.append(["完整路径", str(OUT)])
    ws0.append(["注意", "长正文仍以各 HTML/JS 源文件为准；本表为索引与操作对照。"])
    for cell in ws0[1]:
        cell.font = Font(bold=True)
    autosize_columns(ws0)

    add_sheet(
        wb,
        "页面与流程",
        ["文件", "作用简述"],
        [
            ("index.html", "世界观、玩法；开始调查 → 001"),
            ("001-oa-home.html", "林敏视角 OA；搜索、登录/找回密码、邮箱"),
            ("search.html", "OA 检索结果页，支持 ?keyword="),
            ("002-news-industrial-visit.html", "gameData.newsArticles[002] 动态渲染"),
            ("003-notice-lifeguard-training.html", "gameData.newsArticles[003] 动态渲染"),
            ("004-report-bone-extraction.html", "静态：湖山县区情（与「王安」检索卡片标题不一致）"),
            ("005-news-couple-story.html", "静态：锦旗/张弛报道"),
            ("006-news-doctor-profile.html", "静态：王安人物专访（未接入 OA 检索）"),
            ("007-desk-linlan.html", "林岚受限 OA；搜索；入口 008"),
            ("008-ticket-destruction-appeal.html", "《文件销毁异议申请》驳回详情"),
            ("mailbox.html", "邮箱（林岚/张弛两套内容）"),
            ("009-wuyou-login.html", "可玩登录+找回（page009Engine.js）"),
            ("009.html", "静态占位，无登录逻辑（易与 009 混淆）"),
            ("010-notes-linlan.html", "日记+隐藏搜索词条"),
            ("010.html", "广告邮件静态详情"),
            ("011-partner-juxiang-site.html", "官网；首次进入可触发顶部轻提示「发现隐藏内容!」（无右侧隐藏条）"),
            ("012-desk-zhangchi.html", "张弛线：下载/解压/录音/报案（page012Engine.js）"),
            ("013-ending-report-to-police.html", "结局1；隐藏入口 → 014（hash 口令）"),
            ("014-ending-shadow-archive.html", "封停公告→终端口令→深层归档检索"),
        ],
    )

    add_sheet(
        wb,
        "账号密码密保",
        ["场景", "条件", "结果/说明"],
        [
            ("001/007 OA 登录", "账号 160423，密码 1234Qwer", "linlan；001 上会 markAction loginLinlan"),
            ("001/007 OA 登录", "账号 zhangchi，密码 92TWSL66", "跳转 012；markAction loginZhangchi"),
            ("001 找回密码（弹窗）", "工号 160423 → 答案 2000年3月16日", "显示密码 1234Qwer；链到 007"),
            ("001 登录框验证答案", "2000年3月16日", "提示密码 1234Qwer"),
            ("012 OA 登录", "zhangchi / 92TWSL66", "登录张弛"),
            ("012 OA 登录", "160423 / 1234Qwer", "跳转 001（非留在 012）"),
            ("009 云笔记登录", "linlanlinmin / QWTR10002", "跳转 010 云笔记"),
            ("009 找回密码", "真名 林岚；手机 19905678235；邮箱找回", "wuyou_reset_mail_sent=1；邮箱出现重置邮件"),
            ("010 隐藏日记", "搜索含 张弛 / 密码 / 张弛 密码", "条目正文：password:zhangchi 92TWSL66"),
            ("014 终端", "管理员密码 JXZZ60@WT11", "解锁归档；013 链接 hash 含 passwordJXZZ60@WT11"),
            ("012 虚拟文件", "重命名为 .zip 再解压或压缩王", "page012Engine 扩展名校验"),
        ],
    )

    add_sheet(
        wb,
        "OA检索_有结果",
        ["关键词", "结果数", "目标页面", "检索卡片标题", "摘要"],
        [
            ("林岚", "1", "003-notice-lifeguard-training.html", "P003 - 《关于开展第一届员工水上救生知识培训的通知》", "员工活动公告详情页"),
            ("陈志刚", "1", "002-news-industrial-visit.html", "P002 - 《无限生物高层调研湖山县工业园…》", "公司要闻详情页"),
            ("张文文", "1", "002-news-industrial-visit.html", "同上", "公司要闻详情页"),
            ("陈志刚、张文文", "1", "002-news-industrial-visit.html", "同上", "整串关键词匹配"),
            ("王安", "1", "004-report-bone-extraction.html", "P004 - 《王安技术专报：新一代骨粉提取工艺》", "实际 004 页面为湖山县区情，需统一"),
            ("湖山县", "2", "005 + 002", "P005 区域调查档案；P002 要闻", "两条结果"),
        ],
    )

    add_sheet(
        wb,
        "OA检索_拒绝与提示",
        ["类型", "关键词或情形", "提示文案"],
        [
            ("拒绝", "同种骨", "无查看权限"),
            ("拒绝", "张弛", "无权限查看"),
            ("拒绝", "B-09", "无权限查看"),
            ("系统", "空输入", "请输入关键词。"),
            ("系统", "无匹配", "无搜索结果"),
            ("系统", "重复检索", "该关键词已检索过，以下为已解锁信息。"),
            ("系统", "新线索", "检索成功，发现新线索。"),
        ],
    )

    add_sheet(
        wb,
        "邮箱mailbox",
        ["用户", "要点"],
        [
            ("返回工作台", "zhangchi → 012；否则 → 001"),
            ("林岚-业务", "B-09批次流程记录补充；数据运营部"),
            ("林岚-广告", "VIP 邮件；链 009 无忧云笔记"),
            ("林岚-重置", "临时密码 QWTR10002；需先完成 009 找回"),
            ("张弛", "多封订阅/论文 alert；无外链"),
        ],
    )

    add_sheet(
        wb,
        "012张弛线操作",
        ["类别", "内容"],
        [
            ("禁用点击", "该模块无权限访问 / 该功能无权限 / 已被禁用"),
            ("下载后提示", "模拟下载已完成：右下角出现下载标识。"),
            ("托盘文案", "B-09B_ultrasound_preview.jpg 下载完成"),
            ("解压失败", "无法打开压缩包：文件格式不受支持。"),
            ("解压成功", "压缩包已打开。 / 已使用压缩王.zip 打开：发现 Negotiation_张文文.wav（20 分钟）。"),
            ("重命名", "文件名不能为空。/ 重命名成功…该格式暂不识别。"),
            ("录音标题", "Negotiation_张文文.wav · 声纹片段提取（部分缺失）"),
            ("调查完成", "录音结束后缓冲→「似乎已经找到真相」→自动 013（已取消弹窗）"),
            ("登录失败", "账号或密码错误"),
        ],
    )

    add_sheet(
        wb,
        "012录音逐字稿",
        ["序号", "台词"],
        [
            ("1", "录音内容，呈现一段声波文，有一段是没有"),
            ("2", "张文文——你把自己精子混迹去的事，上面知道了，赵总的手段你是知道的，按照规定，你应该付什么代价你也门清"),
            ("3", "张弛——我没有杀人，充其量只算是医疗失误，况且这里算哪门子医疗部门，货物也是你们销毁的，和我有什么关系？"),
            ("4", "张文文——行啊，你去告，赵总后面是什么，你张弛再不聪明也大概能猜到。"),
            ("5", "一堆混乱的声音，听不清"),
            ("6", "张文文——你自己没管好你们部分的老鼠，解决她，我算你将功赎过"),
        ],
    )

    add_sheet(
        wb,
        "013结局1要点",
        ["项", "内容"],
        [
            ("标题区", "ENDING 01；你拿到录音，去了警局"),
            ("情节关键词", "巨人观、遗书、暗恋林岚、代孕孩子火葬场、无限生物解散、外包简历、公厕小广告"),
            ("广告文案", "挣大钱 女大学生来 有限生物 1889900934"),
            ("按钮", "探索隐藏内容 → 014；返回首页 → 001"),
        ],
    )

    add_sheet(
        wb,
        "014归档",
        ["项", "内容"],
        [
            ("管理员密码", "JXZZ60@WT11"),
            ("URL 锚点标记", "passwordJXZZ60@WT11"),
            ("失败提示", "> AUTHENTICATION FAILED. 请重试。"),
            ("通过提示", "> 校验通过。正在挂载只读卷…… / ACCESS GRANTED. …"),
            ("顶栏", "无线生物科技公司机密文件…"),
            ("文件 cwb-112304", "关键词：cwb,112304,汇款,转账,泰和,银行,回单,牙科,材料,张力,大川,txn9823004128,12800；图 ./media/cwb-112304.png"),
            ("文件 ADM-112306", "关键词：adm,112306,架构,组织,架构图,部门,org,chart；组织架构文本"),
            ("文件 ARC-120627", "arc,120627,辅助生殖…；正文待补充；影响进度第二段"),
            ("文件 LOG-1336890", "log,1336890,物流,运筹；待补充；影响进度第二段"),
            ("文件 00-readme", "readme,read,说明,索引；占位"),
        ],
    )

    add_sheet(
        wb,
        "进度条progressTracker",
        ["阶段", "文案"],
        [
            ("0", "开始调查"),
            ("1", "已打开 003 新闻2"),
            ("2", "已打开 005 页面"),
            ("3", "已登录林岚账号"),
            ("4", "已进入林岚邮箱"),
            ("5", "已进入林岚云笔记"),
            ("6", "已登录张弛账号"),
            ("7", "已修改文件"),
            ("8", "已查看录音文件"),
            ("主轨标题", "探索进度"),
            ("隐藏页 Toast", "发现隐藏内容!（首次进 011 等且主线未满；顶部浮层）"),
            ("已移除 UI 字面", "+40% 隐藏 / 隐藏区域 / 未标注路径…（仅作历史检索，不在界面显示）"),
            ("已移除提示", "系统校验异常：右侧出现未登记进度区间（隐藏关卡）。（代码已删）"),
            ("014 模式", "深层卷宗检索：主线已闭合 · 异常数据段 …"),
        ],
    )

    add_sheet(
        wb,
        "010云笔记",
        ["项", "内容"],
        [
            ("普通日记", "日记1～9：见 010-notes-linlan.html 内 diaries"),
            ("隐藏-简报", "搜索含 A城日报 → 赵宏辉/赵强安 + 链 011"),
            ("隐藏-密码", "搜索含 张弛/密码/张弛 密码 → password:zhangchi 92TWSL66"),
            ("空列表", "没有找到对应日记。"),
            ("隐藏文件夹占位", "日记不可查看，请通过其他方式搜寻"),
        ],
    )

    add_sheet(
        wb,
        "易混与非主线",
        ["文件/模块", "说明"],
        [
            ("gameLogic.js", "仓库/手套/监控证据库；未接入 gameEngine"),
            ("009.html vs 009-wuyou-login.html", "前者无登录脚本"),
            ("page001Engine.js", "DOMContentLoaded 末尾 localStorage.oa_current_user=linlan；与序章林敏设定需核对"),
        ],
    )

    add_sheet(
        wb,
        "验收待统一",
        ["项", "说明"],
        [
            ("公司名", "无限生物 / 无线生物 / 有限生物 混用"),
            ("张弛/张驰", "008 等为「张驰」"),
            ("001 显示名", "界面显示林岚 vs 序章林敏"),
            ("王安检索", "卡片写王安专报，004 实际为湖山县区情；006 才是王安长文"),
            ("013 结局1", "终场段落已替换；公厕广告为「有限生物公司」与正文公司名形成对照"),
            ("012 录音", "声波文 — 是否应为声波纹"),
        ],
    )

    add_sheet(
        wb,
        "002003新闻摘要",
        ["页", "标题/要点"],
        [
            ("002", "《无限生物高层调研湖山县工业园…》；来源 A城日报；陈文/张文；末段「无线生物科技公司」"),
            ("003", "水上救生培训通知；林岚工号 160423 溺水；行政人事部落款"),
        ],
    )

    wb.save(OUT)
    print("Wrote:", OUT)


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
生成「游戏完整文案」Excel：先由 Node 抽取结构化 JSON，再合并各 HTML 可见文本。

运行前提：已安装 openpyxl；本机有 node。
  python _export_full_game_copy_xlsx.py

输出：game_full_copy_inventory.xlsx（与脚本同目录）
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
BUNDLE_JSON = ROOT / "_copy_bundle.json"
NODE_SCRIPT = ROOT / "_export_copy_extract.mjs"
OUT_XLSX = ROOT / "game_full_copy_inventory.xlsx"


def run_node_bundle():
    if not NODE_SCRIPT.is_file():
        raise FileNotFoundError(NODE_SCRIPT)
    r = subprocess.run(
        ["node", str(NODE_SCRIPT)],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if r.returncode != 0:
        sys.stderr.write(r.stderr or "")
        raise RuntimeError("node _export_copy_extract.mjs failed")
    BUNDLE_JSON.write_text(r.stdout, encoding="utf-8")


def load_bundle():
    # 每次生成前重新跑 Node，避免 Windows 下手工重定向产生非 UTF-8 的 _copy_bundle.json
    run_node_bundle()
    return json.loads(BUNDLE_JSON.read_text(encoding="utf-8"))


class VisibleHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self._skip = 0
        self.segments: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style", "noscript"):
            self._skip += 1

    def handle_endtag(self, tag):
        if tag in ("script", "style", "noscript") and self._skip:
            self._skip -= 1

    def handle_data(self, data):
        if self._skip:
            return
        t = re.sub(r"\s+", " ", data).strip()
        if len(t) >= 2:
            self.segments.append(t)


def html_visible_segments(path: Path) -> list[str]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    p = VisibleHTMLParser()
    try:
        p.feed(raw)
        p.close()
    except Exception:
        return [raw[:2000] + ("…" if len(raw) > 2000 else "")]
    return p.segments


def autosize_columns(ws, max_width=96):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = min(max_width, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = w


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize_columns(ws)
    return ws


def flatten_game_data(b):
    gd = b["gameData"]
    rows = []
    for pid, page in (gd.get("pages") or {}).items():
        rows.append(
            (
                f"gameData.js::pages.{pid}.title",
                "gameData_pages",
                "gameData.js",
                f"pages/{pid}/title",
                page.get("title") or "",
                "",
            )
        )
        rows.append(
            (
                f"gameData.js::pages.{pid}.subtitle",
                "gameData_pages",
                "gameData.js",
                f"pages/{pid}/subtitle",
                page.get("subtitle") or "",
                "",
            )
        )
    for aid, art in (gd.get("newsArticles") or {}).items():
        rows.append(
            (
                f"gameData.js::newsArticles.{aid}.title",
                "gameData_news",
                "gameData.js",
                f"newsArticles/{aid}/title",
                art.get("title") or "",
                "",
            )
        )
        rows.append(
            (
                f"gameData.js::newsArticles.{aid}.source",
                "gameData_news",
                "gameData.js",
                f"newsArticles/{aid}/source",
                art.get("source") or "",
                "",
            )
        )
        for bi, block in enumerate(art.get("blocks") or []):
            bt = block.get("type") or ""
            if "text" in block:
                rows.append(
                    (
                        f"gameData.js::newsArticles.{aid}.blocks.{bi}.text",
                        "gameData_news",
                        "gameData.js",
                        f"newsArticles/{aid}/blocks[{bi}]/{bt}",
                        block.get("text") or "",
                        "",
                    )
                )
            if "alt" in block:
                rows.append(
                    (
                        f"gameData.js::newsArticles.{aid}.blocks.{bi}.alt",
                        "gameData_news",
                        "gameData.js",
                        f"newsArticles/{aid}/blocks[{bi}]/alt",
                        block.get("alt") or "",
                        "",
                    )
                )
            if "caption" in block:
                rows.append(
                    (
                        f"gameData.js::newsArticles.{aid}.blocks.{bi}.caption",
                        "gameData_news",
                        "gameData.js",
                        f"newsArticles/{aid}/blocks[{bi}]/caption",
                        block.get("caption") or "",
                        "",
                    )
                )
    for kw, msg in (gd.get("deniedKeywords") or {}).items():
        rows.append(
            (
                f"gameData.js::deniedKeywords.{kw}",
                "gameData_search",
                "gameData.js",
                f"deniedKeywords/{kw}",
                msg,
                f"触发词：{kw}",
            )
        )
    for kw, rec in (gd.get("searchIndex") or {}).items():
        for ri, item in enumerate(rec.get("results") or []):
            rid = item.get("id") or str(ri)
            rows.append(
                (
                    f"gameData.js::searchIndex.{kw}.results.{rid}.title",
                    "gameData_search",
                    "gameData.js",
                    f"searchIndex/{kw}/{rid}/title",
                    item.get("title") or "",
                    "",
                )
            )
            rows.append(
                (
                    f"gameData.js::searchIndex.{kw}.results.{rid}.summary",
                    "gameData_search",
                    "gameData.js",
                    f"searchIndex/{kw}/{rid}/summary",
                    item.get("summary") or "",
                    "",
                )
            )
    return rows


def flatten_mailbox(b):
    rows = []
    for user_key, label in (("linlan", "林岚邮箱"), ("zhangchi", "张弛邮箱")):
        mmap = (b.get("mailbox") or {}).get(
            "linlanMailMap" if user_key == "linlan" else "zhangchiMailMap",
            {},
        )
        for mid, mail in mmap.items():
            rows.append(
                (
                    f"mailboxEngine.js::{user_key}.{mid}.title",
                    "mailbox",
                    "mailboxEngine.js",
                    f"{label}/{mid}/title",
                    mail.get("title") or "",
                    "",
                )
            )
            rows.append(
                (
                    f"mailboxEngine.js::{user_key}.{mid}.meta",
                    "mailbox",
                    "mailboxEngine.js",
                    f"{label}/{mid}/meta",
                    mail.get("meta") or "",
                    "",
                )
            )
            body = mail.get("body") or []
            rows.append(
                (
                    f"mailboxEngine.js::{user_key}.{mid}.body",
                    "mailbox",
                    "mailboxEngine.js",
                    f"{label}/{mid}/body",
                    "\n".join(str(x) for x in body),
                    "多段正文：换行分隔",
                )
            )
            rows.append(
                (
                    f"mailboxEngine.js::{user_key}.{mid}.linkLabel",
                    "mailbox",
                    "mailboxEngine.js",
                    f"{label}/{mid}/linkLabel",
                    mail.get("linkLabel") or "",
                    "",
                )
            )
    return rows


def flatten_archive014(b):
    rows = []
    for item in b.get("archive014") or []:
        fid = item.get("id") or item.get("label")
        rows.append(
            (
                f"014-ending-shadow-archive.html::ARCHIVE_FILES.{fid}.label",
                "archive014",
                "014-ending-shadow-archive.html",
                f"{fid}/label",
                item.get("label") or "",
                "",
            )
        )
        kws = item.get("keywords") or []
        rows.append(
            (
                f"014-ending-shadow-archive.html::ARCHIVE_FILES.{fid}.keywords",
                "archive014",
                "014-ending-shadow-archive.html",
                f"{fid}/keywords",
                ", ".join(str(x) for x in kws),
                "检索关键词，逗号分隔",
            )
        )
        if item.get("bodyText"):
            rows.append(
                (
                    f"014-ending-shadow-archive.html::ARCHIVE_FILES.{fid}.bodyText",
                    "archive014",
                    "014-ending-shadow-archive.html",
                    f"{fid}/bodyText",
                    item.get("bodyText") or "",
                    "",
                )
            )
        if item.get("imageAlt"):
            rows.append(
                (
                    f"014-ending-shadow-archive.html::ARCHIVE_FILES.{fid}.imageAlt",
                    "archive014",
                    "014-ending-shadow-archive.html",
                    f"{fid}/imageAlt",
                    item.get("imageAlt") or "",
                    "",
                )
            )
    return rows


def flatten_diary010(b):
    rows = []
    d10 = b.get("diary010") or {}
    for group_name in ("diaries", "lockedDiaries", "specialDiaries"):
        for i, d in enumerate(d10.get(group_name) or []):
            title = d.get("title") or ""
            rows.append(
                (
                    f"010-notes-linlan.html::{group_name}.{i}.title",
                    "diary010",
                    "010-notes-linlan.html",
                    f"{group_name}[{i}]/title",
                    title,
                    "",
                )
            )
            rows.append(
                (
                    f"010-notes-linlan.html::{group_name}.{i}.date",
                    "diary010",
                    "010-notes-linlan.html",
                    f"{group_name}[{i}]/date",
                    d.get("date") or "",
                    "",
                )
            )
            content = d.get("content") or []
            rows.append(
                (
                    f"010-notes-linlan.html::{group_name}.{i}.content",
                    "diary010",
                    "010-notes-linlan.html",
                    f"{group_name}[{i}]/content",
                    "\n".join(str(x) for x in content),
                    "可含 HTML 片段",
                )
            )
            triggers = d.get("trigger") or d.get("readTriggers")
            if triggers:
                rows.append(
                    (
                        f"010-notes-linlan.html::{group_name}.{i}.triggers",
                        "diary010",
                        "010-notes-linlan.html",
                        f"{group_name}[{i}]/triggers",
                        ", ".join(str(x) for x in triggers),
                        "触发词，勿随意改逻辑关键词除非同步改代码",
                    )
                )
    return rows


def flatten_engine_progress(b):
    rows = []
    for k, v in (b.get("gameEngine") or {}).items():
        rows.append(
            (
                f"gameEngine.js::{k}",
                "gameEngine",
                "gameEngine.js",
                k,
                v,
                "",
            )
        )
    for i, lab in enumerate(b.get("progressLabels") or []):
        rows.append(
            (
                f"progressTracker.js::STEP_LABELS.{i}",
                "progressTracker",
                "progressTracker.js",
                f"STEP_LABELS[{i}]",
                lab,
                "",
            )
        )
    for k, v in (b.get("progressMisc") or {}).items():
        rows.append(
            (
                f"progressTracker.js::misc.{k}",
                "progressTracker",
                "progressTracker.js",
                k,
                v,
                "",
            )
        )
    for i, line in enumerate(b.get("transcript012") or []):
        rows.append(
            (
                f"page012Engine.js::transcriptLines.{i}",
                "page012",
                "page012Engine.js",
                f"录音逐字稿[{i}]",
                line,
                "",
            )
        )
    return rows


def flatten_html_files():
    rows = []
    html_files = sorted(ROOT.glob("*.html"))
    for path in html_files:
        name = path.name
        segs = html_visible_segments(path)
        for si, seg in enumerate(segs):
            rows.append(
                (
                    f"html::{name}::seg_{si:04d}",
                    "html_visible",
                    name,
                    f"可见文本段[{si}]",
                    seg,
                    "由 HTML 解析（不含 script/style）；与内联脚本中的文案可能重复，以 stable_id 为准修改对应源文件",
                )
            )
    return rows


def main():
    b = load_bundle()
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "使用说明"
    now = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M %z")
    readme_rows = [
        ["游戏完整文案表", ""],
        ["生成时间", now],
        ["输出文件", OUT_XLSX.name],
        ["", ""],
        ["列说明", ""],
        ["stable_id", "全局唯一锚点；你改文案后把本表发回，我将按此列写回源码。"],
        ["category", "分组（只读参考）。"],
        ["source_file", "主来源文件名。"],
        ["field", "字段路径（只读参考）。"],
        ["copy_zh", "【在此列修改中文文案】长文本保留换行。"],
        ["notes", "技术说明；一般无需改。"],
        ["", ""],
        ["注意", "1) 「html_visible」与 gameData/内联脚本可能重复，改稿时优先按 stable_id 指向的 source_file 修改。"],
        ["", "2) 谜题账号密码若改字，需同步改对应 .html / .js 中的校验逻辑。"],
        ["", "3) 重新导出会覆盖本文件；请另存为你的副本再编辑。"],
    ]
    for r in readme_rows:
        ws0.append(r)
    for cell in ws0[1]:
        cell.font = Font(bold=True)
    autosize_columns(ws0)

    headers = ["stable_id", "category", "source_file", "field", "copy_zh", "notes"]

    add_sheet(
        wb,
        "gameData_pages_news_search",
        headers,
        flatten_game_data(b),
    )
    add_sheet(wb, "mailbox", headers, flatten_mailbox(b))
    add_sheet(wb, "archive014_diary010", headers, flatten_archive014(b) + flatten_diary010(b))
    add_sheet(wb, "engine_progress_012", headers, flatten_engine_progress(b))
    add_sheet(wb, "html_visible_all", headers, flatten_html_files())

    wb.save(OUT_XLSX)
    print("Wrote:", OUT_XLSX)


if __name__ == "__main__":
    main()
